import requests
from bs4 import BeautifulSoup 
import time 
import openpyxl
import csv
from poisk_elementov import Page 


class Key: 
    def __init__(self, phraze, position):
        self.position = position 
        self.phraze = phraze


class Keys: 
    def __init__(self):
        self.keys = []
    
    def add(self, key): 
        self.keys.append(key)
    def get(self): 
        return self.keys

class Landing_Page: 
    def __init__(self, url, keys): 
        self.url = url 
        self.keys = keys 
    
    def print_keys(self):
        print(self.url)
        for key in self.keys.get(): 
            print(key.position, key.phraze)

    def down_position(self):
        for key in self.keys.get(): 
            if int(key.position) > 10 or int(key.position) == 0: 
                print(key.position, key.phraze)

    def meta(self): 
        if self.url != '': 
            page = Page(self.url)
            page.get_meta()
            page.get_h1()

            self.title = page.title
            self.description = page.description
            self.keywords = page.keywords

            self.h1 = page.h1
        else: 
            self.title = 'Старницы не сущесвтует '
            self.description = 'Старницы не сущесвтует'
            self.keywords = 'Старницы не сущесвтует'
            self.h1 = 'Старницы не сущесвтует'

class LK_Command: 

    # логин и пароль, авторизация 
    def __init__(self, login, password): 
        pass
        self.login = login
        self.password = password
    
        session = requests.Session()
        url = 'http://an.rem-mach.ru/login.php'
        dann = dict(user_login = self.login, user_password = self.password)
        session.get(url)
        session.post(url, dann)

        self.session = session

    def set_site_url(self):
        con = self.session.get("http://an.rem-mach.ru/edit_site_tab_url.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        site_url = html.select('#site_url')
        self.site_url = site_url[0].get('value')
    
    def set_regions(self): 
        con = self.session.get("http://an.rem-mach.ru/keys_list.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        self.regions = html.select('nobr>input')

    def set_price(self): 
        con = self.session.get("http://an.rem-mach.ru/edit_site_tab_budget.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        price= html.select('#site_budg_seo')
        self.price = price[0].get('value')

    def downland_TR(self, number):
        self.number = number
        self.set_site_url() 
        self.set_regions() 
        
        regions = [] 
        for region in self.regions: 
            regions.append(region.get('value')) 

        self.regions_tr = regions 

        for region in regions: 
            self.region = region
            con = self.session.get("http://an.rem-mach.ru/keys_list.php?download_list=1&site_id="+self.number+"&site_stat_region="+str(region))
            f=open(r''+ self.site_url + str(region) + '.csv',"wb")
            f.write(con.content)
            f.close()
    
    def analization_TR(self):
        
        compilation = [] 
        with open(self.site_url+str(self.region) +'.csv', newline='') as File:  
            reader = csv.reader(File)
            for val, row in enumerate(reader):
                if val != 0: 
                    position = row[0].split(';')[3]
                    if position == '-': 
                        position = int(0)
                    
                    url = row[0].split(';')[4]
                    phraze = row[0].split(';')[0]
                    compilation.append({
                        'url': url,
                        'position': position, 
                        'phraze': phraze, 
                    })

        self.compilation = compilation 
        self.group_phraze_in_url()

    # Вывести все слова в топе 
    def get_praze_down(self):
        for lp in self.landing_pages:
            lp.down_position()
        
    def group_phraze_in_url(self):
        mas_pages = []
        self.landing_pages = [] 
        for row in self.compilation: 
            if row['url'] not in mas_pages: 
                mas_pages.append(row['url'])
                url = row['url']
                keys = Keys()  
                for row_2 in self.compilation: 
                    if row_2['url'] == row['url']: 
                        phraze = row_2['phraze']
                        position = row_2['position']
                        key = Key(phraze, position)
                        keys.add(key)
                lp = Landing_Page(url, keys)
                self.landing_pages.append(lp)
    def meta(self): 
        for lp in self.landing_pages: 
            lp.meta() 
            try: 
                print(lp.url)
                print(lp.title) 
                print(lp.description) 
                print(lp.keywords)
            except: 
                pass

    def reported_struct(self, name_file):     
        wb = openpyxl.Workbook()
        sheet = wb.active
        for val, lp in enumerate(self.landing_pages): 
           
            try: 
                cell = sheet.cell(row=val, column=len(lp.url.split('/'))-2)
                cell.hyperlink = lp.url 
                cell.value = lp.h1 
            except: 
                pass 
        wb.save(name_file + ".xlsx")

    def excel(self, mas, name_file): 
        wb = openpyxl.Workbook()
        sheet = wb.active

        for row, product in enumerate(mas): 
            for col, key in enumerate(product.keys()): 
                cell = sheet.cell(row=row+1, column=col+1) 
                cell.value = product[key]
        wb.save(name_file + ".xlsx")


user = LK_Command('Jilcov', 'Jilcov435333')
user.downland_TR('4028')
user.analization_TR()   
user.meta()
user.reported_struct('struct_studiomix') 