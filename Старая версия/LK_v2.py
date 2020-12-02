
import requests
from bs4 import BeautifulSoup 
import time 
import openpyxl
import csv





class LK: 

    def __init__(self, login, password): 
        self.login = login
        self.password = password

    # Авторизация 
    def autorization(self): 
        session = requests.Session()
        url = 'http://an.rem-mach.ru/login.php'
        dann = dict(user_login = self.login, user_password = self.password)
        session.get(url)
        session.post(url, dann)

        self.session = session

    # загрузка подборки в ЛК
    def load_selection(self, url_project, file_name): 
        # работа с файлом
        wb = openpyxl.load_workbook(file_name)
        type(wb)
        sheet_excel = wb.active
        rows_excel  = sheet_excel.max_row

        row_mas = []
        big_data = []

        for row in range(2, rows_excel+1):
            for column in range (1, 6):
                cell = sheet_excel.cell(row = row, column = column)
                row_mas.append(' ' if cell.value == None else cell.value )
            big_data.append(dict({'key': row_mas[0], 
                                'url': row_mas[4], 
                            }))
            row_mas = [] 
        # конец работы с файлом 
        print(big_data)
        # работа в ЛК 
        url_project = url_project
        for s in url_project.split('/'): 
            if s.isdigit(): 
                url_project = "http://an.rem-mach.ru/keys_list.php?site_id="+str(s)
                print(s)

        con = self.session.get(url_project)
        html = BeautifulSoup(con.content, "html.parser")
        id_project_bd = html.select("#advert_site_id") # получение значения id в БД 
        
        #отправка запроса
        for page in big_data: 
            url = 'http://an.rem-mach.ru/keys_list.php?page_in_prop='+page['url']+'&id_in_table='+id_project_bd[0].get('value')+'&key_in_table='+page['key']+'&KEYPOSTPROP=1' 
            print(url)
            self.session.get(url)
        # конец работы в ЛК 

        # обнуление переменных     
        big_data = [] 

    #загрузка релевантныйх страниц 
    def download_selection(self, url_project):

        con = self.session.get(url_project)
        html = BeautifulSoup(con.content, "html.parser")
        site_url_name = html.select("#div_data")
        
        print(site_url_name)

        for s in url_project.split('/'): 
            if s.isdigit(): 
                url_project = "http://an.rem-mach.ru/keys_list.php?site_id="+str(s)
                print(s)

        con = self.session.get(url_project)
        html = BeautifulSoup(con.content, "html.parser")
        site_id = html.select("#site_id")[0]
        site_stat_region = html.select("#site_stat_region")[0]
       
        # url_project = "http://an.rem-mach.ru/keys_list.php?download_list=1&site_id="+site_id.get('value')+"&site_stat_region="+site_stat_region.get('value')
        # url = self.session.get(url_project)
        # f = open(site_url_name.text+".csv", "wb")
        # f.write(url.content) 


    def set_site_url(self):
        con = self.session.get("http://an.rem-mach.ru/edit_site_tab_url.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        site_url = html.select('#site_url')
        self.site_url = site_url[0].get('value')

    def set_regions(self): 
        con = self.session.get("http://an.rem-mach.ru/keys_list.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        self.regions = html.select('nobr>input')

    def set_price(self): 
        con = self.session.get("http://an.rem-mach.ru/edit_site_tab_budget.php?site_id=" + self.number)
        html = BeautifulSoup(con.content, "html.parser")
        price= html.select('#site_budg_seo')
        self.price = price[0].get('value')

    # скачать подборку 
    def count_keyword(self, number):
        # /keys_list.php?download_list=1&site_id=2086&site_stat_region=581
        #site_url , reg , price

        self.number = number

        self.set_site_url() 
        self.set_regions() 
        self.set_price()
        
        reg = [] 
        for region in self.regions: 
            reg.append(region.get('value')) 

        count = 0 

        for r in reg: 
            con = self.session.get("http://an.rem-mach.ru/keys_list.php?download_list=1&site_id="+self.number+"&site_stat_region="+str(r))
            f=open(r''+ self.site_url + str(r) + '.csv',"wb")
            f.write(con.content)
            f.close()

            with open(self.site_url + str(r) +'.csv', newline='') as File:  
                reader = csv.reader(File)
                for row in reader:  
                    for r in row: 
                        stroka = r.split(';')
                        #print(stroka[0], stroka[4])
                        count += 1
                

                print('URL проекта: ', 'http://an.rem-mach.ru/sites/'+self.number +'/')
                print('Название сайта: ', self.site_url)
                print('Регион: ', reg)
                print('Бюджет: ', self.price)
                print('Кол фраз в подборке', count)
                count = 0 

    def analization_report(self, number):
        self.number = number

        self.set_site_url() 
        self.set_regions() 
        self.set_price()
        
        reg = [] 
        for region in self.regions: 
            reg.append(region.get('value')) 

        count = 0 

        for r in reg: 
            con = self.session.get("http://an.rem-mach.ru/keys_list.php?download_list=1&site_id="+self.number+"&site_stat_region="+str(r))
            f=open(r''+ self.site_url + str(r) + '.csv',"wb")
            f.write(con.content)
            f.close()
            
            mas = []
            mas_url = [] 
            with open(self.site_url + str(r) +'.csv', newline='') as File:  
                reader = csv.reader(File)
                for row in reader:
                    for r in row: 
                        
                        stroka = r.split(';')
                        #print(stroka[0], stroka[4])
                        if stroka[3] == 'Яндекс':
                            continue  
                        if stroka[3] == '-':
                            position = int(0)
                        else: position = int(stroka[3])
                        mas.append({'pharaz': stroka[0],
                                     'url': stroka[4],
                                     'position': position 
                                     } 
                                     )
                        count += 1                                       

            self.down_keys = [] 
            copy_mas = [] 
            keys = [] 
            for i in mas: 
                if i['url'] in copy_mas: 
                    pass
                else:
                    url = i['url']
                    copy_mas.append(url)
                    for j in mas: 
                        if j['url'] == i['url']:
                            if j['position'] > 10 or j['position'] == 0: 
                                keys.append(j['pharaz'])
                    self.down_keys.append({'url': url, 'keys': keys.copy()})
                    keys = [] 


    # загрузка тегов в ЛК 
    def load_meta():    
        pass




